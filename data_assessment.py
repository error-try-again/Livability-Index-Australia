import logging
import os
import pandas as pd
from openpyxl import load_workbook
from multiprocessing import Pool, cpu_count

# Configuration constants
COLUMN_NAME_LENGTH_THRESHOLD = 50
NON_NULL_MEAN_THRESHOLD = 0.2
SKIP_ROWS_RANGE = (10, 20)

logging.basicConfig(filename="data_assessment.log", level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')


def is_valid_excel_file(filename):
    if not filename.endswith('.xlsx') or not os.path.exists(filename):
        return False
    try:
        load_workbook(filename, read_only=True)
        return True
    except Exception as e:
        logging.error(f"Error loading {filename}: {e}")
        return False


def extract_excel_metadata(filename):
    wb = load_workbook(filename, read_only=True, data_only=True)
    return {
        key: getattr(wb.properties, key.lower(), "N/A")
        for key in ["Title", "Author", "Description", "Subject", "Identifier",
                    "Version", "Last Modified By", "Category", "Keywords",
                    "Created Date", "Modified Date"]
    }


def check_metadata_integrity(metadata):
    missing_keys = [key for key, value in metadata.items() if value == "N/A"]
    for key in missing_keys:
        logging.warning(f"Document {key} is missing.")


def check_data_integrity(df):
    missing_values = df.isnull().sum().sum()
    if missing_values:
        logging.warning(f"Found {missing_values} missing values in the dataset.")

    duplicate_rows = df.duplicated().sum()
    if duplicate_rows:
        logging.warning(f"Found {duplicate_rows} duplicate rows in the dataset.")


def load_excel_data(filename):
    xls = pd.ExcelFile(filename)
    sheets = xls.sheet_names

    for sheet in sheets:
        df = pd.read_excel(filename, sheet_name=sheet, skiprows=range(*SKIP_ROWS_RANGE))
        df = df.dropna(how='all')
        df.columns = ["Key", "Value"]

        merged_rows = []
        for i, row in df.iterrows():
            key = row['Key']
            value = row['Value']

            if pd.notna(key) and pd.isna(value) and (i + 1) < len(df) and pd.isna(df.iloc[i + 1]['Key']):
                value = df.iloc[i + 1]['Value']

            merged_rows.append({"Key": key, "Value": value})

        df = pd.DataFrame(merged_rows)
        if not df.empty and df.notnull().mean().mean() > NON_NULL_MEAN_THRESHOLD:
            return df
    return None


def save_df_to_json(df, output_dir, filename):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    json_filename = os.path.join(output_dir, f"{filename}.json")
    df.to_json(json_filename, orient='records', date_format='iso')


def process_file(filepath, output_dir):
    logging.info(f"Processing: {filepath}")

    if not is_valid_excel_file(filepath):
        logging.error(f"{filepath} is not a valid Excel file or does not exist.")
        return

    metadata = extract_excel_metadata(filepath)
    check_metadata_integrity(metadata)

    df = load_excel_data(filepath)
    if df is not None:
        check_data_integrity(df)
        save_df_to_json(df, output_dir, os.path.splitext(os.path.basename(filepath))[0])


def process_all_files_in_directory(input_dir, output_dir):
    all_files = [os.path.join(root, file) for root, _, files in os.walk(input_dir) for file in files if file.endswith('.xlsx')]
    with Pool(processes=cpu_count()) as pool:
        pool.starmap(process_file, [(file, output_dir) for file in all_files])


if __name__ == "__main__":
    INPUT_DIRECTORY = "./test"
    OUTPUT_DIRECTORY = "./output"
    process_all_files_in_directory(INPUT_DIRECTORY, OUTPUT_DIRECTORY)
